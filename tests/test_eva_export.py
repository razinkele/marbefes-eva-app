"""
Tests for eva_export.py — workbook generation, NaN handling, and error resilience.
"""

import sys
import os
from unittest.mock import patch

import numpy as np
import openpyxl
import pandas as pd
import pytest

# Ensure project root is on sys.path
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))

from eva_export import build_workbook


# ---------------------------------------------------------------------------
# Fixtures / helpers
# ---------------------------------------------------------------------------

def _minimal_inputs(n_subzones=3, n_features=2, inject_nan=False):
    """Return a valid set of arguments for build_workbook."""
    data = {"Subzone ID": [f"SZ_{i+1}" for i in range(n_subzones)]}
    for j in range(n_features):
        col_vals = list(np.random.default_rng(42 + j).uniform(0, 10, n_subzones))
        if inject_nan and j == 0:
            col_vals[0] = np.nan
        data[f"Feature_{j+1}"] = col_vals

    df = pd.DataFrame(data)

    results = pd.DataFrame({
        "Subzone ID": df["Subzone ID"],
        "AQ1": [1.0, np.nan, 3.0] if inject_nan else [1.0, 2.0, 3.0],
        "AQ2": [4.0, 5.0, np.nan] if inject_nan else [4.0, 5.0, 6.0],
        "EV": [2.5, np.nan, 4.5] if inject_nan else [2.5, 3.5, 4.5],
    })

    classifications = {f"Feature_{j+1}": [] for j in range(n_features)}
    metadata = {
        "ec_name": "TestEC",
        "study_area": "TestArea",
        "data_description": "Test data",
    }
    ec_store = {
        "TestEC": {
            "data": df,
            "data_type": "quantitative",
            "classifications": classifications,
            "results": results,
            "feature_count": n_features,
        }
    }
    return dict(
        results=results,
        uploaded_data=df,
        user_classifications=classifications,
        data_type="quantitative",
        metadata=metadata,
        ec_store=ec_store,
    )


# ---------------------------------------------------------------------------
# Tests
# ---------------------------------------------------------------------------

class TestBuildWorkbook:

    def test_returns_workbook(self):
        """build_workbook returns an openpyxl Workbook with expected sheets."""
        inputs = _minimal_inputs()
        wb = build_workbook(**inputs)
        assert isinstance(wb, openpyxl.Workbook)
        assert len(wb.sheetnames) > 0

    def test_null_returns_info_sheet(self):
        """None results returns a minimal workbook with an Info sheet."""
        inputs = _minimal_inputs()
        inputs["results"] = None
        wb = build_workbook(**inputs)
        assert isinstance(wb, openpyxl.Workbook)
        assert "Info" in wb.sheetnames
        # Check that the info message is present
        ws = wb["Info"]
        assert ws.cell(row=2, column=1).value == "No data available"

    def test_aq_results_nan_preserved(self):
        """NaN values in AQ/EV columns appear as None (empty cells), NOT 0."""
        inputs = _minimal_inputs(inject_nan=True)
        wb = build_workbook(**inputs)

        # Check "AQ & EV Results" sheet
        ws_aq = wb["AQ & EV Results"]
        # Read all data values from the sheet (skip header row 1)
        values = []
        for row in ws_aq.iter_rows(min_row=2, values_only=True):
            values.extend(row)

        # NaN should become None (empty cell), never 0
        # We injected NaN in AQ1 row 2, AQ2 row 3, EV row 2
        # Verify that None appears (empty cells exist)
        assert None in values, "Expected empty cells (None) for NaN values"

        # Verify no unexpected 0 replaced the NaN — specifically check the
        # cells where we injected NaN.
        # Row 2 (data row index 0→excel row 2): AQ1=1.0, AQ2=4.0, EV=2.5
        # Row 3 (data row index 1→excel row 3): AQ1=NaN, AQ2=5.0, EV=NaN
        # Row 4 (data row index 2→excel row 4): AQ1=3.0, AQ2=NaN, EV=4.5
        aq1_row3 = ws_aq.cell(row=3, column=2).value  # AQ1 for SZ_2
        ev_row3 = ws_aq.cell(row=3, column=4).value    # EV for SZ_2
        aq2_row4 = ws_aq.cell(row=4, column=3).value   # AQ2 for SZ_3

        assert aq1_row3 is None, f"AQ1 NaN should be empty cell, got {aq1_row3}"
        assert ev_row3 is None, f"EV NaN should be empty cell, got {ev_row3}"
        assert aq2_row4 is None, f"AQ2 NaN should be empty cell, got {aq2_row4}"

        # Also check "Complete Results" sheet
        ws_complete = wb["Complete Results"]
        complete_values = []
        for row in ws_complete.iter_rows(min_row=2, values_only=True):
            complete_values.extend(row)
        assert None in complete_values, (
            "Complete Results should also have empty cells for NaN"
        )

    def test_has_expected_sheets(self):
        """Workbook contains all expected sheet names for a single-EC run."""
        inputs = _minimal_inputs()
        wb = build_workbook(**inputs)
        expected = {
            "Summary & Metadata",
            "Original Data",
            "AQ & EV Results",
            "Feature Classifications",
            "AQ Methodology",
            "EV Calculation",
            "Complete Results",
        }
        actual = set(wb.sheetnames)
        assert expected.issubset(actual), (
            f"Missing sheets: {expected - actual}"
        )

    @patch("eva_export.pio.to_image", side_effect=RuntimeError("kaleido missing"))
    def test_chart_failure_handled(self, mock_to_image):
        """Workbook is still valid even when chart generation fails."""
        inputs = _minimal_inputs()
        wb = build_workbook(**inputs)
        assert isinstance(wb, openpyxl.Workbook)
        # Core data sheets should still be present
        assert "AQ & EV Results" in wb.sheetnames
        assert "Complete Results" in wb.sheetnames
        # Chart error sheet should exist
        assert "Chart Errors" in wb.sheetnames

    @patch("eva_export.pio.to_image", side_effect=RuntimeError("kaleido missing"))
    def test_multi_ec_uses_max_aggregation(self, _mock_img):
        """Multi-EC Total EV must use MAX, not SUM."""
        from eva_config import ECEntry

        # Create two ECs with known EV values per subzone
        subzones = ["SZ_1", "SZ_2", "SZ_3"]
        df_a = pd.DataFrame({
            "Subzone ID": subzones,
            "Feature_A": [1.0, 2.0, 3.0],
        })
        results_a = pd.DataFrame({
            "Subzone ID": subzones,
            "AQ1": [2.0, 3.0, 4.0],
            "EV": [2.0, 3.0, 4.0],
        })
        df_b = pd.DataFrame({
            "Subzone ID": subzones,
            "Feature_B": [5.0, 6.0, 7.0],
        })
        results_b = pd.DataFrame({
            "Subzone ID": subzones,
            "AQ1": [3.0, 1.0, 5.0],
            "EV": [3.0, 1.0, 5.0],
        })

        ec_store = {
            "EC_Alpha": ECEntry(
                data=df_a,
                data_type="quantitative",
                classifications={"Feature_A": []},
                results=results_a,
            ),
            "EC_Beta": ECEntry(
                data=df_b,
                data_type="quantitative",
                classifications={"Feature_B": []},
                results=results_b,
            ),
        }

        metadata = {
            "ec_name": "Multi",
            "study_area": "TestArea",
            "data_description": "Multi-EC test",
        }

        wb = build_workbook(
            results=results_a,
            uploaded_data=df_a,
            user_classifications={"Feature_A": []},
            data_type="quantitative",
            metadata=metadata,
            ec_store=ec_store,
        )

        # The "Aggregated EV" sheet should exist
        assert "Aggregated EV" in wb.sheetnames, (
            f"Expected 'Aggregated EV' sheet, got: {wb.sheetnames}"
        )

        ws = wb["Aggregated EV"]
        # Data starts at row 3 (row 1 = title, row 3 = header)
        # Find the "Total EV" column index
        header_row = 3
        total_ev_col = None
        for col_idx in range(1, ws.max_column + 1):
            if ws.cell(row=header_row, column=col_idx).value == "Total EV":
                total_ev_col = col_idx
                break
        assert total_ev_col is not None, "Could not find 'Total EV' column"

        # Find Subzone ID column
        sz_col = None
        for col_idx in range(1, ws.max_column + 1):
            if ws.cell(row=header_row, column=col_idx).value == "Subzone ID":
                sz_col = col_idx
                break
        assert sz_col is not None, "Could not find 'Subzone ID' column"

        # Read all Total EV values keyed by subzone
        total_ev_by_sz = {}
        for row_idx in range(header_row + 1, ws.max_row + 1):
            sz = ws.cell(row=row_idx, column=sz_col).value
            ev = ws.cell(row=row_idx, column=total_ev_col).value
            if sz is not None:
                total_ev_by_sz[sz] = ev

        # Expected: MAX(EC_Alpha, EC_Beta) per subzone
        # SZ_1: max(2.0, 3.0) = 3.0
        # SZ_2: max(3.0, 1.0) = 3.0
        # SZ_3: max(4.0, 5.0) = 5.0
        assert total_ev_by_sz["SZ_1"] == pytest.approx(3.0), (
            f"SZ_1 Total EV should be MAX(2,3)=3.0, got {total_ev_by_sz['SZ_1']}"
        )
        assert total_ev_by_sz["SZ_2"] == pytest.approx(3.0), (
            f"SZ_2 Total EV should be MAX(3,1)=3.0, got {total_ev_by_sz['SZ_2']}"
        )
        assert total_ev_by_sz["SZ_3"] == pytest.approx(5.0), (
            f"SZ_3 Total EV should be MAX(4,5)=5.0, got {total_ev_by_sz['SZ_3']}"
        )
