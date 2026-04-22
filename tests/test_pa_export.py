"""
Tests for pa_export.py -- standalone PA workbook, combined workbook, and BBT8 workbook.
"""

import io
import sys
import os
from unittest.mock import patch, MagicMock

import numpy as np
import openpyxl
import pandas as pd
import pytest

# Ensure project root is on sys.path
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))

from pa_export import (
    _build_extent_sheet,
    generate_bbt8_workbook,
    generate_combined_workbook,
    generate_pa_workbook,
)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _minimal_extent_df():
    """Return a small extent DataFrame with a EUNIS code column and area."""
    return pd.DataFrame({
        "eunis_code": ["MA1", "MB1", "MC1"],
        "area_ha": [100.0, 250.5, 75.0],
    })


def _minimal_supply_df():
    """Return a small supply DataFrame."""
    return pd.DataFrame({
        "EUNIS_code": ["MA1", "MB1"],
        "service_1": [0.8, 0.6],
        "service_2": [0.3, 0.9],
    })


def _minimal_assignments():
    return {"SZ_1": "MA1", "SZ_2": "MB1", "SZ_3": "MC1"}


def _minimal_metadata():
    return {
        "eaa_name": "Test EAA",
        "boundary_description": "Test boundary",
        "accounting_year": "2025",
    }


def _minimal_bbt8_inputs():
    """Return minimal DataFrames for generate_bbt8_workbook."""
    accounts = pd.DataFrame({
        "EUNIS_code": ["MA1", "MB1"],
        "EUNIS_name": ["Littoral rock", "Infralittoral rock"],
        "total_area": [100.0, 250.0],
        "mean_EV": [3.5, 4.0],
        "confidence": [0.8, 0.9],
    })
    main_values = pd.DataFrame({
        "Subzone_ID": ["SZ_1", "SZ_2", "SZ_3"],
        "EUNIS_code": ["MA1", "MB1", "MA1"],
        "Habitat_EV": [3.0, 4.0, 4.0],
        "Habitat_confidence": [0.8, 0.9, 0.7],
    })
    extent = pd.DataFrame({
        "EUNIS_code": ["MA1", "MB1"],
        "total_area": [100.0, 250.0],
    })
    condition = pd.DataFrame({
        "EUNIS_code": ["MA1", "MB1"],
        "mean_condition": [0.75, 0.85],
    })
    supply = pd.DataFrame({
        "EUNIS_code": ["MA1", "MB1"],
        "service_proxy": [0.6, 0.8],
    })
    metadata = {
        "EAA": "Test EAA",
        "Year": "2025",
        "Description": "BBT8 test",
    }
    return dict(
        accounts=accounts,
        main_values=main_values,
        extent=extent,
        condition=condition,
        supply=supply,
        metadata=metadata,
    )


# ---------------------------------------------------------------------------
# Tests: generate_pa_workbook
# ---------------------------------------------------------------------------

class TestGeneratePAWorkbook:

    def test_returns_bytesio(self):
        """generate_pa_workbook returns a BytesIO buffer."""
        buf = generate_pa_workbook(
            extent_df=_minimal_extent_df(),
            supply_df=_minimal_supply_df(),
            assignments=_minimal_assignments(),
            metadata=_minimal_metadata(),
            completeness="100%",
            unit="Ha",
        )
        assert isinstance(buf, io.BytesIO)
        # Buffer should contain a valid xlsx
        buf.seek(0)
        wb = openpyxl.load_workbook(buf)
        assert len(wb.sheetnames) > 0

    def test_has_expected_sheets(self):
        """PA workbook contains all 5 expected sheets."""
        buf = generate_pa_workbook(
            extent_df=_minimal_extent_df(),
            supply_df=_minimal_supply_df(),
            assignments=_minimal_assignments(),
            metadata=_minimal_metadata(),
            completeness="85%",
        )
        wb = openpyxl.load_workbook(buf)
        expected = {
            "Summary & Metadata",
            "Ecosystem Extent Account",
            "Supply Table",
            "Habitat Assignments",
            "Methodology",
        }
        actual = set(wb.sheetnames)
        assert expected.issubset(actual), f"Missing sheets: {expected - actual}"

    def test_extent_sheet_has_totals_row(self):
        """Ecosystem Extent Account sheet has a TOTAL row at the bottom."""
        buf = generate_pa_workbook(
            extent_df=_minimal_extent_df(),
            supply_df=None,
            assignments=None,
            metadata=_minimal_metadata(),
            completeness="100%",
        )
        wb = openpyxl.load_workbook(buf)
        ws = wb["Ecosystem Extent Account"]
        last_row = ws.max_row
        assert ws.cell(row=last_row, column=1).value == "TOTAL"

    def test_none_extent_handled(self):
        """None extent_df produces a valid workbook with 'No data' placeholder."""
        buf = generate_pa_workbook(
            extent_df=None,
            supply_df=None,
            assignments=None,
            metadata=_minimal_metadata(),
            completeness="0%",
        )
        wb = openpyxl.load_workbook(buf)
        ws = wb["Ecosystem Extent Account"]
        assert ws.cell(row=2, column=1).value == "No data available"

    def test_none_supply_handled(self):
        """None supply_df produces a 'No supply data' placeholder."""
        buf = generate_pa_workbook(
            extent_df=_minimal_extent_df(),
            supply_df=None,
            assignments=_minimal_assignments(),
            metadata=_minimal_metadata(),
            completeness="50%",
        )
        wb = openpyxl.load_workbook(buf)
        ws = wb["Supply Table"]
        assert ws.cell(row=2, column=1).value == "No supply data available"

    def test_empty_assignments_handled(self):
        """Empty assignments dict produces a placeholder row."""
        buf = generate_pa_workbook(
            extent_df=_minimal_extent_df(),
            supply_df=_minimal_supply_df(),
            assignments={},
            metadata=_minimal_metadata(),
            completeness="100%",
        )
        wb = openpyxl.load_workbook(buf)
        ws = wb["Habitat Assignments"]
        assert ws.cell(row=2, column=1).value == "No assignments available"

    def test_summary_contains_metadata_fields(self):
        """Summary sheet contains the EAA name from metadata."""
        buf = generate_pa_workbook(
            extent_df=_minimal_extent_df(),
            supply_df=None,
            assignments=None,
            metadata=_minimal_metadata(),
            completeness="100%",
        )
        wb = openpyxl.load_workbook(buf)
        ws = wb["Summary & Metadata"]
        # Scan for the EAA name value
        values = [ws.cell(row=r, column=2).value for r in range(1, ws.max_row + 1)]
        assert "Test EAA" in values


# ---------------------------------------------------------------------------
# Tests: generate_combined_workbook
# ---------------------------------------------------------------------------

class TestGenerateCombinedWorkbook:

    def _make_eva_args(self):
        """Build minimal eva_args dict for build_workbook."""
        df = pd.DataFrame({
            "Subzone ID": ["SZ_1", "SZ_2"],
            "Feature_1": [1.0, 2.0],
        })
        results = pd.DataFrame({
            "Subzone ID": ["SZ_1", "SZ_2"],
            "AQ1": [1.0, 2.0],
            "EV": [2.5, 3.5],
        })
        return dict(
            results=results,
            uploaded_data=df,
            user_classifications={"Feature_1": []},
            data_type="quantitative",
            metadata={
                "ec_name": "TestEC",
                "study_area": "TestArea",
                "data_description": "Test",
            },
            ec_store={
                "TestEC": {
                    "data": df,
                    "data_type": "quantitative",
                    "classifications": {"Feature_1": []},
                    "results": results,
                    "feature_count": 1,
                }
            },
        )

    @patch("eva_export.pio.to_image", side_effect=RuntimeError("no kaleido"))
    def test_returns_bytesio(self, _mock_img):
        """generate_combined_workbook returns a BytesIO buffer."""
        buf = generate_combined_workbook(
            eva_args=self._make_eva_args(),
            pa_extent_df=_minimal_extent_df(),
            pa_supply_df=_minimal_supply_df(),
            pa_assignments=_minimal_assignments(),
            pa_metadata=_minimal_metadata(),
            pa_completeness="100%",
            pa_unit="Ha",
        )
        assert isinstance(buf, io.BytesIO)
        buf.seek(0)
        wb = openpyxl.load_workbook(buf)
        assert len(wb.sheetnames) > 0

    @patch("eva_export.pio.to_image", side_effect=RuntimeError("no kaleido"))
    def test_has_pa_sheets(self, _mock_img):
        """Combined workbook contains PA-prefixed sheets."""
        buf = generate_combined_workbook(
            eva_args=self._make_eva_args(),
            pa_extent_df=_minimal_extent_df(),
            pa_supply_df=_minimal_supply_df(),
            pa_assignments=_minimal_assignments(),
            pa_metadata=_minimal_metadata(),
            pa_completeness="100%",
        )
        wb = openpyxl.load_workbook(buf)
        pa_sheets = {
            "PA - Extent Account",
            "PA - Supply Table",
            "PA - Habitat Assignments",
            "PA - Methodology",
        }
        actual = set(wb.sheetnames)
        assert pa_sheets.issubset(actual), f"Missing PA sheets: {pa_sheets - actual}"


# ---------------------------------------------------------------------------
# Tests: generate_bbt8_workbook
# ---------------------------------------------------------------------------

class TestGenerateBBT8Workbook:

    def test_returns_bytesio(self):
        """generate_bbt8_workbook returns a BytesIO buffer."""
        inputs = _minimal_bbt8_inputs()
        buf = generate_bbt8_workbook(**inputs)
        assert isinstance(buf, io.BytesIO)

    def test_has_expected_sheets(self):
        """BBT8 workbook contains the standard sheet set."""
        inputs = _minimal_bbt8_inputs()
        buf = generate_bbt8_workbook(**inputs)
        wb = openpyxl.load_workbook(buf)
        expected = {
            "ReadMe",
            "main_values",
            "habitat_area_sum",
            "accounts",
            "condition",
            "supply",
        }
        actual = set(wb.sheetnames)
        assert expected.issubset(actual), f"Missing sheets: {expected - actual}"

    def test_missing_values_sheet_omitted_when_none(self):
        """missing_values sheet is NOT created when missing_values=None."""
        inputs = _minimal_bbt8_inputs()
        inputs["missing_values"] = None
        buf = generate_bbt8_workbook(**inputs)
        wb = openpyxl.load_workbook(buf)
        assert "missing_values" not in wb.sheetnames

    def test_missing_values_sheet_present_when_provided(self):
        """missing_values sheet IS created when a non-empty DataFrame is given."""
        inputs = _minimal_bbt8_inputs()
        inputs["missing_values"] = pd.DataFrame({
            "EUNIS_code": ["MC1"],
            "gap_type": ["no_data"],
        })
        buf = generate_bbt8_workbook(**inputs)
        wb = openpyxl.load_workbook(buf)
        assert "missing_values" in wb.sheetnames

    def test_readme_contains_metadata(self):
        """ReadMe sheet contains the metadata key-value pairs."""
        inputs = _minimal_bbt8_inputs()
        buf = generate_bbt8_workbook(**inputs)
        wb = openpyxl.load_workbook(buf)
        ws = wb["ReadMe"]
        # Row 2 should have first metadata key/value (row 1 is header)
        assert ws.cell(row=2, column=1).value == "EAA"
        assert ws.cell(row=2, column=2).value == "Test EAA"

    def test_habitat_area_sum_uses_total_area(self):
        """When extent has 'total_area' (not 'area_m2'), habitat_area_sum uses it."""
        inputs = _minimal_bbt8_inputs()
        buf = generate_bbt8_workbook(**inputs)
        wb = openpyxl.load_workbook(buf)
        ws = wb["habitat_area_sum"]
        # Header row
        assert ws.cell(row=1, column=1).value == "EUNIS2019C"
        assert ws.cell(row=1, column=2).value == "Sum of area"


# ---------------------------------------------------------------------------
# TestBuildExtentSheet
# ---------------------------------------------------------------------------

class TestBuildExtentSheet:
    def test_custom_habitat_name_preserved(self):
        """Custom habitat names (not in EUNIS_LOOKUP) must survive export."""
        df = pd.DataFrame({
            "eunis_code":   ["X99"],
            "habitat_name": ["Custom reef mosaic"],
            "area":         [42.0],
            "pct_total":    [100.0],
        })
        wb = openpyxl.Workbook()
        ws = wb.active
        _build_extent_sheet(ws, df, unit="Ha")
        data_row = list(ws.iter_rows(min_row=2, max_row=2, values_only=True))[0]
        # Columns: EUNIS Code, Habitat Name, Area (Ha), % of Total
        assert data_row[0] == "X99"
        assert data_row[1] == "Custom reef mosaic", (
            f"Expected custom name, got {data_row[1]!r}"
        )

    def test_empty_habitat_name_falls_back_to_lookup(self):
        """If habitat_name is empty/NaN for a real EUNIS code, fall back to lookup."""
        df = pd.DataFrame({
            "eunis_code":   ["MB252", "MB252"],
            "habitat_name": ["", None],  # both empty — CSV round-trip scenarios
            "area":         [10.0, 20.0],
            "pct_total":    [33.3, 66.7],
        })
        wb = openpyxl.Workbook()
        ws = wb.active
        _build_extent_sheet(ws, df, unit="Ha")
        rows = list(ws.iter_rows(min_row=2, max_row=3, values_only=True))
        # Both rows should show the real EUNIS_LOOKUP name, not empty string
        assert rows[0][1] == "Posidonia oceanica meadows", (
            f"Empty-string habitat_name should fall back to EUNIS_LOOKUP, got {rows[0][1]!r}"
        )
        assert rows[1][1] == "Posidonia oceanica meadows", (
            f"None habitat_name should fall back to EUNIS_LOOKUP, got {rows[1][1]!r}"
        )

    def test_pct_total_from_dataframe_is_preserved(self):
        """If DataFrame carries pct_total, exporter must use it — not recompute.

        Fixture choice: areas 1.0 and 2.0 would yield recomputed percentages
        of 33.33 and 66.67. We deliberately override pct_total to 50.0/50.0
        so the test fails before the fix and passes after.
        """
        df = pd.DataFrame({
            "eunis_code":   ["A",   "B"],
            "habitat_name": ["Alpha", "Beta"],
            "area":         [1.0,   2.0],
            "pct_total":    [50.0,  50.0],  # override, NOT matching area ratio
        })
        wb = openpyxl.Workbook()
        ws = wb.active
        _build_extent_sheet(ws, df, unit="Ha")
        rows = list(ws.iter_rows(min_row=2, max_row=3, values_only=True))
        # % column is index 3
        assert rows[0][3] == 50.0, f"Expected 50.0 (from pct_total), got {rows[0][3]}"
        assert rows[1][3] == 50.0, f"Expected 50.0 (from pct_total), got {rows[1][3]}"

    def test_total_row_reflects_actual_pct_sum(self):
        """TOTAL row % must equal sum of pct_total, not the hardcoded 100.0."""
        df = pd.DataFrame({
            "eunis_code":   ["A",   "B"],
            "habitat_name": ["Alpha", "Beta"],
            "area":         [1.0,   1.0],
            "pct_total":    [40.0,  40.0],  # partial coverage — sums to 80
        })
        wb = openpyxl.Workbook()
        ws = wb.active
        _build_extent_sheet(ws, df, unit="Ha")
        # TOTAL row is row 4 (header=1, A=2, B=3, TOTAL=4)
        total_row = list(ws.iter_rows(min_row=4, max_row=4, values_only=True))[0]
        assert total_row[0] == "TOTAL"
        assert total_row[3] == 80.0, (
            f"Expected TOTAL row % = 80.0 (sum of pct_total), got {total_row[3]}"
        )


# ---------------------------------------------------------------------------
# TestGenerateBbt8WorkbookSchema
# ---------------------------------------------------------------------------

class TestGenerateBbt8WorkbookSchema:
    def _valid_inputs(self):
        """Build a minimal set of valid inputs; individual tests mutate one."""
        extent = pd.DataFrame({"EUNIS_code": ["MB252"], "area_m2": [100.0]})
        accounts = pd.DataFrame({
            "EUNIS_code": ["MB252"], "EUNIS_name": ["Posidonia"],
            "area_m2": [100.0], "Habitat_EV": [0.5], "Confidence": [0.9],
        })
        main_values = pd.DataFrame({
            "Subzone_ID": ["A"], "EUNIS_code": ["MB252"],
            "Habitat_EV": [0.5], "Habitat_confidence": [0.9],
        })
        condition = pd.DataFrame({"EUNIS_code": ["MB252"]})
        supply = pd.DataFrame({"EUNIS_code": ["MB252"]})
        return extent, accounts, main_values, condition, supply

    def test_missing_extent_area_column_raises(self):
        """extent without area_m2 or total_area must raise ValueError."""
        extent, accounts, main_values, condition, supply = self._valid_inputs()
        extent = pd.DataFrame({"EUNIS_code": ["MB252"], "something_else": [1]})
        with pytest.raises(ValueError, match="area_m2"):
            generate_bbt8_workbook(
                accounts=accounts, main_values=main_values, extent=extent,
                condition=condition, supply=supply, metadata={"BBT": "test"},
            )

    def test_missing_main_values_subzone_id_raises(self):
        """main_values without Subzone_ID must raise ValueError."""
        extent, accounts, main_values, condition, supply = self._valid_inputs()
        main_values = main_values.drop(columns=["Subzone_ID"])
        with pytest.raises(ValueError, match="Subzone_ID"):
            generate_bbt8_workbook(
                accounts=accounts, main_values=main_values, extent=extent,
                condition=condition, supply=supply, metadata={"BBT": "test"},
            )

    def test_valid_inputs_do_not_raise(self):
        """Sanity: valid inputs must not trigger the schema guard."""
        extent, accounts, main_values, condition, supply = self._valid_inputs()
        generate_bbt8_workbook(
            accounts=accounts, main_values=main_values, extent=extent,
            condition=condition, supply=supply, metadata={"BBT": "test"},
        )
