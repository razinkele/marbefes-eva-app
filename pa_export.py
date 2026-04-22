"""
MARBEFES Physical Accounts (PA) Excel Export Module

Generates Excel workbooks for Physical Accounts — both standalone and
combined with EVA results.  All functions are stateless and have no
Shiny dependencies.  Accepts plain Python objects and returns an
io.BytesIO buffer.
"""

import io
import logging

import numpy as np
import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter

import eva_export
from eva_export import style_worksheet
from pa_config import (
    PA_MODULE_VERSION,
    PA_METHODOLOGY,
    EXPORT_PA_TAB_COLOR,
    EUNIS_LOOKUP,
)

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Private helpers
# ---------------------------------------------------------------------------

def _build_pa_summary_sheet(ws, metadata, extent_df, completeness):
    """Write PA summary rows to a worksheet.

    Writes Parameter/Value pairs covering module name, version,
    export date/time, EAA name, boundary description, accounting year,
    data completeness, habitat count, total extent, and references.

    Parameters
    ----------
    ws : openpyxl.worksheet.worksheet.Worksheet
        Target worksheet (should be empty).
    metadata : dict
        Keys: ``eaa_name``, ``boundary_description``, ``accounting_year``.
        All keys are optional; missing ones are rendered as empty strings.
    extent_df : pd.DataFrame or None
        DataFrame with at least one numeric area column used to derive
        habitat count and total extent.  May be None or empty.
    completeness : str or float
        Data completeness indicator (e.g. ``"100%"`` or ``0.85``).
    """
    # Derive statistics from extent_df
    habitat_count = 0
    total_extent = 0.0
    if extent_df is not None and not (
        isinstance(extent_df, pd.DataFrame) and extent_df.empty
    ):
        df = extent_df if isinstance(extent_df, pd.DataFrame) else pd.DataFrame(extent_df)
        habitat_count = len(df)
        # Look up area column by name (robust to column-order changes)
        area_col = next(
            (c for c in ["area", "total_area", "area_Ha", "area_km2"] if c in df.columns),
            None,
        )
        if area_col is None:
            numeric_cols = df.select_dtypes(include=[np.number]).columns.tolist()
            area_col = numeric_cols[0] if numeric_cols else None
        if area_col:
            total_extent = float(df[area_col].sum())

    now = pd.Timestamp.now()
    rows = [
        ("Parameter", "Value"),
        ("Module", "Physical Accounts (PA)"),
        ("PA Module Version", PA_MODULE_VERSION),
        ("Export Date", now.strftime("%Y-%m-%d")),
        ("Export Time", now.strftime("%H:%M:%S")),
        ("EAA Name", metadata.get("eaa_name", "")),
        ("Boundary Description", metadata.get("boundary_description", "")),
        ("Accounting Year", metadata.get("accounting_year", "")),
        ("Data Completeness", f"{completeness.get('pct', 0):.1f}% ({completeness.get('filled', 0)}/{completeness.get('total', 0)} cells)" if isinstance(completeness, dict) else str(completeness)),
        ("Habitat Count", habitat_count),
        ("Total Extent", f"{total_extent:,.4f}"),
        ("", ""),
        ("Reference", "Franco A. and Amorim E. (2025) Ecological Value Assessment (EVA)"),
        ("Funding", "European Union Horizon Europe Research Programme - MARBEFES Project"),
    ]

    for row in rows:
        ws.append(list(row))


def _build_extent_sheet(ws, extent_df, unit):
    """Write Ecosystem Extent Account data to a worksheet.

    Headers: EUNIS Code, Habitat Name, Area(<unit>), % of Total.
    A totals row is appended at the bottom.

    Parameters
    ----------
    ws : openpyxl.worksheet.worksheet.Worksheet
        Target worksheet (should be empty).
    extent_df : pd.DataFrame or None
        DataFrame expected to have columns ``eunis_code`` (or similar) and
        one numeric area column.  Column detection is best-effort.
    unit : str
        Area unit label used in the header (e.g. ``"Ha"`` or ``"km2"``).
    """
    if extent_df is None or (isinstance(extent_df, pd.DataFrame) and extent_df.empty):
        ws.append(["EUNIS Code", f"Habitat Name", f"Area ({unit})", "% of Total"])
        ws.append(["No data available", "", "", ""])
        return

    df = extent_df.copy() if isinstance(extent_df, pd.DataFrame) else pd.DataFrame(extent_df)

    # Detect EUNIS code column (case-insensitive, partial match)
    code_col = None
    for col in df.columns:
        if "eunis" in col.lower() or "code" in col.lower() or "habitat" in col.lower():
            code_col = col
            break
    if code_col is None:
        code_col = df.columns[0]

    # Detect area column (first numeric column that is not the code column)
    numeric_cols = df.select_dtypes(include=[np.number]).columns.tolist()
    area_col = numeric_cols[0] if numeric_cols else None

    # Write header
    ws.append(["EUNIS Code", "Habitat Name", f"Area ({unit})", "% of Total"])

    total_area = float(df[area_col].sum()) if area_col else 0.0

    for _, row in df.iterrows():
        code = str(row[code_col]) if code_col else ""
        if "habitat_name" in df.columns:
            name_raw = row["habitat_name"]
            if pd.isna(name_raw) or not str(name_raw).strip():
                name = EUNIS_LOOKUP.get(code, "")
            else:
                name = str(name_raw)
        else:
            name = EUNIS_LOOKUP.get(code, "")
        area = float(row[area_col]) if area_col else 0.0
        pct = (area / total_area * 100) if total_area > 0 else 0.0
        ws.append([code, name, round(area, 4), round(pct, 2)])

    # Totals row
    ws.append(["TOTAL", "", round(total_area, 4), 100.0])


def _build_supply_sheet(ws, supply_df):
    """Write Supply Table DataFrame to a worksheet.

    Handles None or empty gracefully by writing a placeholder message.

    Parameters
    ----------
    ws : openpyxl.worksheet.worksheet.Worksheet
        Target worksheet (should be empty).
    supply_df : pd.DataFrame or None
        Supply table data.  If None or empty, a placeholder is written.
    """
    if supply_df is None or (isinstance(supply_df, pd.DataFrame) and supply_df.empty):
        ws.append(["Info"])
        ws.append(["No supply data available"])
        return

    df = supply_df if isinstance(supply_df, pd.DataFrame) else pd.DataFrame(supply_df)

    # Write header row
    ws.append(list(df.columns))

    # Write data rows
    for _, row in df.iterrows():
        ws.append([
            None if (isinstance(v, float) and np.isnan(v)) else v
            for v in row.tolist()
        ])


def _build_assignments_sheet(ws, assignments):
    """Write habitat assignment mapping to a worksheet.

    Writes a header row followed by one row per entry in the
    ``{subzone_id: eunis_code}`` mapping, with the habitat name resolved
    from EUNIS_LOOKUP.

    Parameters
    ----------
    ws : openpyxl.worksheet.worksheet.Worksheet
        Target worksheet (should be empty).
    assignments : dict or None
        Mapping of ``{subzone_id: eunis_code}``.  None or empty dict
        results in a placeholder row.
    """
    ws.append(["Subzone ID", "EUNIS Code", "Habitat Name"])

    if not assignments:
        ws.append(["No assignments available", "", ""])
        return

    for subzone_id, eunis_code in assignments.items():
        habitat_name = EUNIS_LOOKUP.get(str(eunis_code), "")
        ws.append([str(subzone_id), str(eunis_code), habitat_name])


def _build_methodology_sheet(ws):
    """Write PA_METHODOLOGY dict to a worksheet.

    PA_METHODOLOGY is a dict of equal-length lists (column name → values).
    Each key becomes a column header; values are written row by row.

    Parameters
    ----------
    ws : openpyxl.worksheet.worksheet.Worksheet
        Target worksheet (should be empty).
    """
    if not PA_METHODOLOGY:
        ws.append(["No methodology data available"])
        return

    headers = list(PA_METHODOLOGY.keys())
    ws.append(headers)

    # Determine row count from first column
    row_count = len(next(iter(PA_METHODOLOGY.values())))
    for i in range(row_count):
        row = [PA_METHODOLOGY[h][i] if i < len(PA_METHODOLOGY[h]) else "" for h in headers]
        ws.append(row)


# ---------------------------------------------------------------------------
# Public functions
# ---------------------------------------------------------------------------

def generate_pa_workbook(
    extent_df,
    supply_df,
    assignments,
    metadata,
    completeness,
    unit="Ha",
):
    """Generate a standalone Physical Accounts Excel workbook.

    Creates a workbook with five sheets:

    1. Summary & Metadata
    2. Ecosystem Extent Account
    3. Supply Table
    4. Habitat Assignments
    5. Methodology

    Each sheet receives ``style_worksheet()`` styling and the PA tab color.

    Parameters
    ----------
    extent_df : pd.DataFrame or None
        Ecosystem extent data.
    supply_df : pd.DataFrame or None
        Ecosystem service supply table.
    assignments : dict or None
        ``{subzone_id: eunis_code}`` mapping.
    metadata : dict
        Keys: ``eaa_name``, ``boundary_description``, ``accounting_year``.
    completeness : str or float
        Data completeness indicator.
    unit : str, optional
        Area unit label (default ``"Ha"``).

    Returns
    -------
    io.BytesIO
        Excel workbook serialised to a BytesIO buffer, seeked to position 0.
    """
    wb = openpyxl.Workbook()

    # Remove the default empty sheet created by openpyxl
    default_sheet = wb.active
    wb.remove(default_sheet)

    # Sheet definitions: (title, builder_callable)
    sheet_specs = [
        (
            "Summary & Metadata",
            lambda ws: _build_pa_summary_sheet(ws, metadata, extent_df, completeness),
        ),
        (
            "Ecosystem Extent Account",
            lambda ws: _build_extent_sheet(ws, extent_df, unit),
        ),
        (
            "Supply Table",
            lambda ws: _build_supply_sheet(ws, supply_df),
        ),
        (
            "Habitat Assignments",
            lambda ws: _build_assignments_sheet(ws, assignments),
        ),
        (
            "Methodology",
            lambda ws: _build_methodology_sheet(ws),
        ),
    ]

    export_errors = []
    for sheet_title, builder in sheet_specs:
        ws = wb.create_sheet(title=sheet_title)
        ws.sheet_properties.tabColor = EXPORT_PA_TAB_COLOR
        try:
            builder(ws)
        except Exception as e:
            logger.exception("Error building PA sheet '%s'", sheet_title)
            export_errors.append(f"{sheet_title}: {e}")
        style_worksheet(ws)

    if export_errors:
        ws_err = wb.create_sheet("Export Errors")
        ws_err.cell(row=1, column=1, value="Sheet Generation Errors")
        ws_err.cell(row=2, column=1, value="The following sheets could not be generated:")
        for i, err in enumerate(export_errors, start=3):
            ws_err.cell(row=i, column=1, value=err)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def generate_combined_workbook(
    eva_args,
    pa_extent_df,
    pa_supply_df,
    pa_assignments,
    pa_metadata,
    pa_completeness,
    pa_unit="Ha",
):
    """Generate a combined EVA + Physical Accounts Excel workbook.

    Calls ``eva_export.build_workbook(**eva_args)`` to obtain the EVA
    workbook, then appends four PA sheets with a ``"PA - "`` prefix:

    * PA - Extent Account
    * PA - Supply Table
    * PA - Habitat Assignments
    * PA - Methodology

    Parameters
    ----------
    eva_args : dict
        Keyword arguments forwarded verbatim to ``eva_export.build_workbook()``.
    pa_extent_df : pd.DataFrame or None
        Ecosystem extent data.
    pa_supply_df : pd.DataFrame or None
        Ecosystem service supply table.
    pa_assignments : dict or None
        ``{subzone_id: eunis_code}`` mapping.
    pa_metadata : dict
        Keys: ``eaa_name``, ``boundary_description``, ``accounting_year``.
    pa_completeness : str or float
        Data completeness indicator.
    pa_unit : str, optional
        Area unit label (default ``"Ha"``).

    Returns
    -------
    io.BytesIO
        Combined Excel workbook serialised to a BytesIO buffer, seeked to
        position 0.
    """
    # Build the EVA workbook (returns openpyxl.Workbook)
    wb = eva_export.build_workbook(**eva_args)

    # PA sheet definitions: (title_suffix, builder_callable)
    pa_sheet_specs = [
        (
            "PA - Extent Account",
            lambda ws: _build_extent_sheet(ws, pa_extent_df, pa_unit),
        ),
        (
            "PA - Supply Table",
            lambda ws: _build_supply_sheet(ws, pa_supply_df),
        ),
        (
            "PA - Habitat Assignments",
            lambda ws: _build_assignments_sheet(ws, pa_assignments),
        ),
        (
            "PA - Methodology",
            lambda ws: _build_methodology_sheet(ws),
        ),
    ]

    export_errors = []
    for sheet_title, builder in pa_sheet_specs:
        ws = wb.create_sheet(title=sheet_title)
        ws.sheet_properties.tabColor = EXPORT_PA_TAB_COLOR
        try:
            builder(ws)
        except Exception as e:
            logger.exception("Error building combined PA sheet '%s'", sheet_title)
            export_errors.append(f"{sheet_title}: {e}")
        style_worksheet(ws)

    if export_errors:
        ws_err = wb.create_sheet("Export Errors")
        ws_err.cell(row=1, column=1, value="Sheet Generation Errors")
        ws_err.cell(row=2, column=1, value="The following sheets could not be generated:")
        for i, err in enumerate(export_errors, start=3):
            ws_err.cell(row=i, column=1, value=err)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def generate_bbt8_workbook(accounts, main_values, extent, condition,
                           supply, metadata, missing_values=None):
    """Generate BBT8-format Excel workbook for Physical Accounts.

    Follows the Irish Sea BBT8 standard with sheets:
    ReadMe, main_values, habitat_area_sum, accounts, missing_values,
    condition, supply.

    Parameters
    ----------
    accounts : pd.DataFrame
        Summary table with EUNIS code, name, area, EV, confidence.
    main_values : pd.DataFrame
        Per-subzone data: Subzone_ID, EUNIS_code, Habitat_EV,
        Habitat_confidence.
    extent : pd.DataFrame
        Area per EUNIS class from ``compute_eunis_extent()``.
    condition : pd.DataFrame
        Condition stats per EUNIS class from ``compute_eunis_condition()``.
    supply : pd.DataFrame
        Ecosystem service proxies from ``compute_eunis_supply()``.
    metadata : dict
        Key-value pairs written to the ReadMe sheet.
    missing_values : pd.DataFrame or None, optional
        Data gaps table.  Omitted from the workbook when None or empty.

    Returns
    -------
    io.BytesIO
        Excel workbook serialised to a BytesIO buffer, seeked to position 0.
    """
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        # ReadMe
        readme_rows = [(k, v) for k, v in metadata.items()]
        pd.DataFrame(readme_rows, columns=["Parameter", "Value"]).to_excel(
            writer, sheet_name="ReadMe", index=False)

        # main_values (per-subzone)
        main_values.to_excel(writer, sheet_name="main_values", index=False)

        # habitat_area_sum
        if "area_m2" in extent.columns:
            area_sum = extent[["EUNIS_code", "area_m2"]].copy()
            area_sum.columns = ["EUNIS2019C", "Sum of area_m2"]
        else:
            area_sum = extent[["EUNIS_code", "total_area"]].copy()
            area_sum.columns = ["EUNIS2019C", "Sum of area"]
        area_sum.to_excel(writer, sheet_name="habitat_area_sum", index=False)

        # accounts (main summary)
        acct = accounts.copy()
        acct.columns = [c.replace("EUNIS_code", "EUNIS2019C").replace("EUNIS_name", "Habitat")
                        for c in acct.columns]
        acct.to_excel(writer, sheet_name="accounts", index=False)

        # missing_values
        if missing_values is not None and not missing_values.empty:
            missing_values.to_excel(writer, sheet_name="missing_values", index=False)

        # condition
        condition.to_excel(writer, sheet_name="condition", index=False)

        # supply
        supply.to_excel(writer, sheet_name="supply", index=False)

    # Apply styling
    buffer.seek(0)
    wb = openpyxl.load_workbook(buffer)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        style_worksheet(ws)
        ws.sheet_properties.tabColor = "006994"
    styled_buffer = io.BytesIO()
    wb.save(styled_buffer)
    styled_buffer.seek(0)
    return styled_buffer
