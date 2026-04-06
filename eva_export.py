"""
MARBEFES EVA Excel Export — workbook generation and styling.

All functions are stateless and have no Shiny dependencies.
Accepts plain Python objects and returns an io.BytesIO buffer.
"""

import io
import logging

import numpy as np
import openpyxl
import pandas as pd
import plotly.graph_objects as go
import plotly.io as pio
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.drawing.image import Image as XlImage

from eva_config import (
    AQ_METHODOLOGY,
    EV_EXPLANATION,
    APP_VERSION,
    EXPORT_TAB_COLORS,
    EXPORT_HEADER_COLOR,
    EXPORT_ALT_ROW_COLOR,
    EXPORT_MULTI_EC_TAB_COLOR,
    EXPORT_CHART_TAB_COLOR,
    HEATMAP_HEIGHT_PER_ROW,
    HEATMAP_MIN_HEIGHT,
    CHART_EXPORT_WIDTH,
    CHART_EXPORT_HEIGHT,
)

from eva_calculations import merge_multi_ec_ev

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Module-level style objects
# ---------------------------------------------------------------------------
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_HEADER_FILL = PatternFill(
    start_color=EXPORT_HEADER_COLOR,
    end_color=EXPORT_HEADER_COLOR,
    fill_type="solid",
)
_HEADER_ALIGNMENT = Alignment(
    horizontal="center", vertical="center", wrap_text=True
)
_THIN_BORDER = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)
_ALT_ROW_FILL = PatternFill(
    start_color=EXPORT_ALT_ROW_COLOR,
    end_color=EXPORT_ALT_ROW_COLOR,
    fill_type="solid",
)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def style_worksheet(ws, has_data=True, freeze=True, autofilter=True, start_row=1):
    """Apply professional styling to a worksheet."""
    if ws.max_row < start_row or ws.max_column < 1:
        return

    # Header row styling
    for cell in ws[start_row]:
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGNMENT
        cell.border = _THIN_BORDER

    if has_data and ws.max_row > start_row:
        # Autofilter
        if autofilter:
            ws.auto_filter.ref = (
                f"A{start_row}:{get_column_letter(ws.max_column)}{ws.max_row}"
            )

        # Freeze panes below header
        if freeze:
            ws.freeze_panes = f"A{start_row + 1}"

        # Data rows: borders + alternating fill
        for row_idx in range(start_row + 1, ws.max_row + 1):
            for cell in ws[row_idx]:
                cell.border = _THIN_BORDER
                if (row_idx - start_row) % 2 == 0:
                    cell.fill = _ALT_ROW_FILL

    # Auto-size columns (estimate from content)
    for col_idx in range(1, ws.max_column + 1):
        max_len = 0
        col_letter = get_column_letter(col_idx)
        for row_idx in range(start_row, min(ws.max_row + 1, start_row + 50)):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(max_len + 3, 10), 60)


def _build_summary_sheet(writer, results, df, data_type, metadata, ec_store):
    """Write the Summary & Metadata sheet."""
    if len(ec_store) >= 2:
        # Multi-EC summary
        summary_rows = [
            ("Analysis Date", pd.Timestamp.now().strftime("%Y-%m-%d")),
            ("Analysis Time", pd.Timestamp.now().strftime("%H:%M:%S")),
            ("Application Version", APP_VERSION),
            ("Study Area", metadata["study_area"]),
            ("Data Description", metadata["data_description"]),
            ("", ""),
            ("Multi-EC Analysis", ""),
            ("Number of ECs", len(ec_store)),
            (
                "Total Features (all ECs)",
                sum(ec["feature_count"] for ec in ec_store.values()),
            ),
        ]

        ev_vals = []
        for ec in ec_store.values():
            if ec["results"] is not None:
                ev_vals.extend(ec["results"]["EV"].tolist())
        if ev_vals:
            # Use MAX-aggregated Total EV per subzone for summary stats (EVA methodology)
            from eva_calculations import merge_multi_ec_ev
            merged_ev = merge_multi_ec_ev(ec_store)
            if merged_ev is not None and "Total EV" in merged_ev.columns:
                total_ev_series = merged_ev["Total EV"]
                summary_rows.extend([
                    ("Mean Total EV (MAX per subzone)", f"{total_ev_series.mean():.4f}"),
                    ("Maximum Total EV", f"{total_ev_series.max():.4f}"),
                    ("Minimum Total EV", f"{total_ev_series.min():.4f}"),
                ])
            else:
                summary_rows.extend([
                    ("Average EV (across all subzones)", f"{np.mean(ev_vals):.4f}"),
                    ("Maximum EV", f"{max(ev_vals):.4f}"),
                    ("Minimum EV", f"{min(ev_vals):.4f}"),
                ])

        summary_rows.extend([
            ("", ""),
            ("Per-EC Details", ""),
        ])

        for ec_name_s, ec in ec_store.items():
            mean_ev = ec["results"]["EV"].mean() if ec["results"] is not None else 0
            max_ev = ec["results"]["EV"].max() if ec["results"] is not None else 0
            summary_rows.append((
                f"  EC: {ec_name_s}",
                f"{ec['data_type']}, {ec['feature_count']} features, "
                f"Mean EV={mean_ev:.2f}, Max EV={max_ev:.2f}",
            ))

        summary_rows.extend([
            ("", ""),
            ("Reference", "Franco A. and Amorim E. (2025) Ecological Value Assessment (EVA)"),
            ("Funding", "European Union Horizon Europe Research Programme - MARBEFES Project"),
        ])

        summary_df = pd.DataFrame(summary_rows, columns=["Parameter", "Value"])
    else:
        # Single-EC summary
        summary_data = {
            "Parameter": [
                "Analysis Date", "Analysis Time", "Application Version",
                "EC Name", "Study Area", "Data Type", "Data Description",
                "", "Dataset Statistics", "Number of Subzones", "Number of Features",
                "Total EV (Sum)", "Average EV", "Maximum EV", "Minimum EV",
                "", "Reference", "Funding",
            ],
            "Value": [
                pd.Timestamp.now().strftime("%Y-%m-%d"),
                pd.Timestamp.now().strftime("%H:%M:%S"),
                APP_VERSION,
                metadata["ec_name"],
                metadata["study_area"],
                data_type if data_type else "Not specified",
                metadata["data_description"],
                "", "",
                len(results),
                len([col for col in df.columns if col != "Subzone ID"]),
                f"{results['EV'].sum():.4f}",
                f"{results['EV'].mean():.4f}",
                f"{results['EV'].max():.4f}",
                f"{results['EV'].min():.4f}",
                "",
                "Franco A. and Amorim E. (2025) Ecological Value Assessment (EVA)",
                "European Union Horizon Europe Research Programme - MARBEFES Project",
            ],
        }
        summary_df = pd.DataFrame(summary_data)

    summary_df.to_excel(writer, sheet_name="Summary & Metadata", index=False)


def _build_data_sheets(writer, results, df, user_classifications):
    """Write Original Data, AQ & EV Results, Feature Classifications,
    AQ Methodology, EV Calculation, and Complete Results sheets."""

    # Sheet 2: Original Data (NaN exports as empty cells)
    df_export = df.copy()
    df_export.to_excel(writer, sheet_name="Original Data", index=False)

    # Sheet 3: Assessment Questions Results (NaN exports as empty cells)
    aq_cols = (
        ["Subzone ID"]
        + [col for col in results.columns if col.startswith("AQ")]
        + ["EV"]
    )
    results_export = results[aq_cols].copy()
    results_export.to_excel(writer, sheet_name="AQ & EV Results", index=False)

    # Sheet 4: Feature Classifications
    if user_classifications:
        feature_cols = [col for col in df.columns if col != "Subzone ID"]
        classifications_data = []
        for feature in feature_cols:
            user_class = user_classifications.get(feature, [])
            classifications_data.append({
                "Feature Name": feature,
                "RRF (Regionally Rare)": "Yes" if "RRF" in user_class else "No",
                "NRF (Nationally Rare)": "Yes" if "NRF" in user_class else "No",
                "ESF (Ecologically Significant)": "Yes" if "ESF" in user_class else "No",
                "HFS/BH (Habitat Forming)": "Yes" if "HFS_BH" in user_class else "No",
                "SS (Symbiotic Species)": "Yes" if "SS" in user_class else "No",
            })
        classifications_df = pd.DataFrame(classifications_data)
        classifications_df.to_excel(
            writer, sheet_name="Feature Classifications", index=False
        )

    # Sheet 5: AQ Methodology Reference
    methodology_df = pd.DataFrame(AQ_METHODOLOGY)
    methodology_df.to_excel(writer, sheet_name="AQ Methodology", index=False)

    # Sheet 6: EV Calculation Explanation
    ev_df = pd.DataFrame(EV_EXPLANATION)
    ev_df.to_excel(writer, sheet_name="EV Calculation", index=False)

    # Sheet 7: Complete Results (NaN exports as empty cells)
    results_complete = results.copy()
    results_complete.to_excel(writer, sheet_name="Complete Results", index=False)


def _build_multi_ec_sheets(writer, results, ec_store):
    """Write Aggregated EV and per-EC result sheets when multiple ECs exist."""
    if len(ec_store) < 2:
        return

    # Aggregation sheet
    merged = merge_multi_ec_ev(ec_store)

    if merged is not None:
        merged = merged.sort_values("Total EV", ascending=False)
        merged.to_excel(
            writer, sheet_name="Aggregated EV", index=False, startrow=2
        )
        ws = writer.sheets["Aggregated EV"]
        ws.cell(row=1, column=1, value="Aggregated Ecological Value Across All ECs")

    # Per-EC result sheets
    for ec_name, ec in ec_store.items():
        if ec["results"] is not None:
            sheet_name = f"EC - {ec_name}"[:31]  # Excel 31-char limit
            ec["results"].to_excel(
                writer, sheet_name=sheet_name, index=False, startrow=2
            )
            ws = writer.sheets[sheet_name]
            ws.cell(
                row=1, column=1,
                value=f"Results for EC: {ec_name} ({ec['data_type']})",
            )


def _build_chart_sheets(workbook, results, ec_store):
    """Create embedded chart sheets (EV bar, AQ heatmap, EV distribution).

    Each chart is wrapped in its own try/except so one failure does not
    prevent the others.  Failed charts produce a placeholder sheet explaining
    the problem instead of silently disappearing.
    """
    chart_errors = []

    # Chart 1: EV by Subzone bar chart
    try:
        if len(ec_store) >= 2:
            merged_chart = merge_multi_ec_ev(ec_store)
            if merged_chart is not None:
                chart_ev_x = merged_chart["Subzone ID"]
                chart_ev_y = merged_chart["Total EV"]
                chart_ev_title = "Total EV by Subzone (Aggregated)"
            else:
                chart_ev_x = results["Subzone ID"]
                chart_ev_y = results["EV"]
                chart_ev_title = "EV by Subzone"
        else:
            chart_ev_x = results["Subzone ID"]
            chart_ev_y = results["EV"]
            chart_ev_title = "EV by Subzone"

        fig_ev = go.Figure(data=[go.Bar(
            x=chart_ev_x,
            y=chart_ev_y,
            marker=dict(
                color=chart_ev_y.tolist(), colorscale="Viridis", showscale=True
            ),
        )])
        fig_ev.update_layout(
            title=chart_ev_title,
            xaxis_title="Subzone ID",
            yaxis_title="EV Score",
            height=CHART_EXPORT_HEIGHT,
            width=CHART_EXPORT_WIDTH,
            plot_bgcolor="rgba(0,0,0,0)",
        )
        ev_img_bytes = pio.to_image(fig_ev, format="png", width=CHART_EXPORT_WIDTH, height=CHART_EXPORT_HEIGHT, scale=2)
        ev_img_stream = io.BytesIO(ev_img_bytes)
        ws_chart1 = workbook.create_sheet("Chart - EV by Subzone")
        ws_chart1.sheet_properties.tabColor = EXPORT_CHART_TAB_COLOR
        img1 = XlImage(ev_img_stream)
        img1.width = CHART_EXPORT_WIDTH
        img1.height = CHART_EXPORT_HEIGHT
        ws_chart1.add_image(img1, "A1")
    except Exception as e:
        logger.warning("EV bar chart failed: %s", e)
        chart_errors.append(f"EV by Subzone: {e}")

    # Chart 2: AQ Heatmap
    try:
        aq_columns = [col for col in results.columns if col.startswith("AQ")]
        if aq_columns:
            display_cols = aq_columns + ["EV"]
            sorted_res = results.sort_values("EV", ascending=True)
            z_data = sorted_res[display_cols].fillna(0).values

            fig_heatmap = go.Figure(data=go.Heatmap(
                z=z_data,
                x=display_cols,
                y=sorted_res["Subzone ID"].tolist(),
                colorscale="Viridis",
                zmin=0,
                zmax=5,
                text=np.round(z_data, 1),
                texttemplate="%{text}",
                textfont={"size": 9},
                colorbar=dict(title="Score"),
            ))
            fig_heatmap.update_layout(
                title="AQ Scores x Subzones (sorted by EV)",
                xaxis_title="Assessment Questions",
                yaxis_title="Subzone ID",
                height=max(HEATMAP_MIN_HEIGHT, len(sorted_res) * HEATMAP_HEIGHT_PER_ROW),
                width=CHART_EXPORT_WIDTH,
                plot_bgcolor="rgba(0,0,0,0)",
            )
            hm_height = max(HEATMAP_MIN_HEIGHT, len(sorted_res) * HEATMAP_HEIGHT_PER_ROW)
            hm_img_bytes = pio.to_image(
                fig_heatmap, format="png", width=CHART_EXPORT_WIDTH, height=hm_height, scale=2
            )
            hm_img_stream = io.BytesIO(hm_img_bytes)
            ws_chart2 = workbook.create_sheet("Chart - AQ Heatmap")
            ws_chart2.sheet_properties.tabColor = EXPORT_CHART_TAB_COLOR
            img2 = XlImage(hm_img_stream)
            img2.width = CHART_EXPORT_WIDTH
            img2.height = hm_height
            ws_chart2.add_image(img2, "A1")
    except Exception as e:
        logger.warning("AQ heatmap chart failed: %s", e)
        chart_errors.append(f"AQ Heatmap: {e}")

    # Chart 3: EV Distribution histogram
    try:
        fig_hist = go.Figure(data=[go.Histogram(
            x=results["EV"],
            nbinsx=20,
            marker=dict(color="#006994", line=dict(color="white", width=1)),
        )])
        fig_hist.update_layout(
            title="EV Score Distribution",
            xaxis_title="EV Score",
            yaxis_title="Count",
            height=CHART_EXPORT_HEIGHT,
            width=CHART_EXPORT_WIDTH,
            plot_bgcolor="rgba(0,0,0,0)",
            bargap=0.05,
        )
        hist_img_bytes = pio.to_image(
            fig_hist, format="png", width=CHART_EXPORT_WIDTH, height=CHART_EXPORT_HEIGHT, scale=2
        )
        hist_img_stream = io.BytesIO(hist_img_bytes)
        ws_chart3 = workbook.create_sheet("Chart - EV Distribution")
        ws_chart3.sheet_properties.tabColor = EXPORT_CHART_TAB_COLOR
        img3 = XlImage(hist_img_stream)
        img3.width = CHART_EXPORT_WIDTH
        img3.height = CHART_EXPORT_HEIGHT
        ws_chart3.add_image(img3, "A1")
    except Exception as e:
        logger.warning("EV distribution chart failed: %s", e)
        chart_errors.append(f"EV Distribution: {e}")

    # If any charts failed, add a summary sheet explaining what happened
    if chart_errors:
        ws_errors = workbook.create_sheet("Chart Errors")
        ws_errors.cell(row=1, column=1, value="Chart Generation Errors")
        ws_errors.cell(row=2, column=1, value="The following charts could not be generated:")
        for i, err in enumerate(chart_errors, start=3):
            ws_errors.cell(row=i, column=1, value=err)
        ws_errors.cell(row=len(chart_errors) + 4, column=1,
                       value="Tip: Ensure kaleido is installed (pip install kaleido)")


def _build_eunis_ev_sheet(writer, results, eunis_overlay_data):
    """Write EV aggregated by EUNIS habitat type.

    Args:
        writer: ExcelWriter
        results: DataFrame with Subzone ID, AQ columns, EV
        eunis_overlay_data: DataFrame with Subzone_ID, dominant_EUNIS, dominant_EUNIS_name
    """
    if eunis_overlay_data is None or eunis_overlay_data.empty:
        return

    # Normalize ID column
    id_col = "Subzone_ID" if "Subzone_ID" in eunis_overlay_data.columns else "Subzone ID"
    res_id_col = "Subzone ID" if "Subzone ID" in results.columns else "Subzone_ID"

    eunis_sub = eunis_overlay_data[[id_col, "dominant_EUNIS", "dominant_EUNIS_name"]].copy()
    eunis_sub = eunis_sub.rename(columns={id_col: "merge_id"})

    res_sub = results.copy()
    res_sub = res_sub.rename(columns={res_id_col: "merge_id"})

    merged = eunis_sub.merge(res_sub, on="merge_id", how="inner")
    if merged.empty:
        return

    # Get AQ and EV columns
    aq_cols = [c for c in merged.columns if c.startswith("AQ") or c == "EV"]

    # Aggregate: mean per EUNIS class
    agg = merged.groupby(["dominant_EUNIS", "dominant_EUNIS_name"])[aq_cols].agg(
        ["mean", "min", "max", "count"]
    ).round(3)

    # Flatten multi-index columns
    agg.columns = [f"{c[0]}_{c[1]}" for c in agg.columns]
    agg = agg.reset_index()
    agg = agg.rename(columns={"dominant_EUNIS": "EUNIS Code", "dominant_EUNIS_name": "Habitat"})

    agg.to_excel(writer, sheet_name="EV by Habitat Type", index=False, startrow=2)
    ws = writer.sheets["EV by Habitat Type"]
    ws.cell(row=1, column=1, value="Ecological Value by EUNIS Habitat Type")


def _apply_styling(workbook):
    """Apply professional styling, conditional formatting, and tab colors
    to every sheet in the workbook."""

    # Apply tab colors and base styling to all sheets
    for sheet_name in workbook.sheetnames:
        ws = workbook[sheet_name]
        if sheet_name in EXPORT_TAB_COLORS:
            ws.sheet_properties.tabColor = EXPORT_TAB_COLORS[sheet_name]

        # Multi-EC sheets have data starting at row 3 (title in row 1, header in row 3)
        if sheet_name.startswith("EC - ") or sheet_name == "Aggregated EV":
            ws.sheet_properties.tabColor = EXPORT_MULTI_EC_TAB_COLOR
            style_worksheet(ws, start_row=3)
        elif sheet_name == "EV by Habitat Type":
            style_worksheet(ws, start_row=3)
        else:
            style_worksheet(ws)

    # Conditional formatting: color scale on EV columns
    ev_sheets = ["AQ & EV Results", "Complete Results"]
    if "Aggregated EV" in workbook.sheetnames:
        ev_sheets.append("Aggregated EV")

    for sheet_name in ev_sheets:
        if sheet_name not in workbook.sheetnames:
            continue
        ws = workbook[sheet_name]
        start_row = 3 if sheet_name == "Aggregated EV" else 1

        for col_idx in range(1, ws.max_column + 1):
            header_cell = ws.cell(row=start_row, column=col_idx)
            if header_cell.value and ("EV" in str(header_cell.value)):
                col_letter = get_column_letter(col_idx)
                data_range = f"{col_letter}{start_row + 1}:{col_letter}{ws.max_row}"
                ws.conditional_formatting.add(
                    data_range,
                    ColorScaleRule(
                        start_type="num", start_value=0, start_color="F8696B",
                        mid_type="num", mid_value=2.5, mid_color="FFEB84",
                        end_type="num", end_value=5, end_color="63BE7B",
                    ),
                )

                for row_idx in range(start_row + 1, ws.max_row + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.number_format = "0.00"

    # Format AQ columns as 2 decimal places
    for sheet_name in ["AQ & EV Results", "Complete Results"]:
        if sheet_name not in workbook.sheetnames:
            continue
        ws = workbook[sheet_name]
        for col_idx in range(1, ws.max_column + 1):
            header_cell = ws.cell(row=1, column=col_idx)
            if header_cell.value and str(header_cell.value).startswith("AQ"):
                for row_idx in range(2, ws.max_row + 1):
                    ws.cell(row=row_idx, column=col_idx).number_format = "0.00"


# ---------------------------------------------------------------------------
# Public entry points
# ---------------------------------------------------------------------------

def build_workbook(results, uploaded_data, user_classifications,
                   data_type, metadata, ec_store, pa_summary_data=None):
    """
    Build and return a complete Excel workbook with all analysis results.

    Parameters
    ----------
    results : pd.DataFrame or None
        DataFrame with AQ scores and EV values.
    uploaded_data : pd.DataFrame or None
        Original uploaded DataFrame.
    user_classifications : dict
        Mapping of ``{feature_name: [classification_list]}``.
    data_type : str
        ``"qualitative"`` or ``"quantitative"``.
    metadata : dict
        Keys: ``ec_name``, ``study_area``, ``data_description``.
    ec_store : dict
        Mapping of ``{ec_name: {data, data_type, classifications,
        results, feature_count}}``.
    pa_summary_data : object, optional
        Physical Accounts summary data for combined PA export (reserved for
        future use by the PA module).

    Returns
    -------
    openpyxl.Workbook
        The finished workbook object.
    """
    # Handle null case — return a minimal Workbook instead of BytesIO
    if results is None or uploaded_data is None:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Info"
        ws.append(["Message"])
        ws.append(["No data available"])
        return wb

    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        # Sheet 1: Summary & Metadata
        _build_summary_sheet(writer, results, uploaded_data, data_type,
                             metadata, ec_store)

        # Sheets 2-7: data, classifications, methodology, EV explanation,
        #             and complete results
        _build_data_sheets(writer, results, uploaded_data, user_classifications)

        # Multi-EC sheets (Aggregated EV + per-EC results)
        _build_multi_ec_sheets(writer, results, ec_store)

        # Embedded chart sheets
        _build_chart_sheets(writer.book, results, ec_store)

        # EUNIS habitat summary (if overlay data provided)
        if pa_summary_data is not None:
            try:
                _build_eunis_ev_sheet(writer, results, pa_summary_data)
            except Exception as e:
                logger.warning("EUNIS EV sheet failed: %s", e)

        # Professional styling, conditional formatting, tab colors
        _apply_styling(writer.book)

    buffer.seek(0)
    return openpyxl.load_workbook(buffer)


def generate_workbook(results, uploaded_data, user_classifications,
                      data_type, metadata, ec_store, pa_summary_data=None):
    """Backward-compatible entry point — returns io.BytesIO buffer."""
    wb = build_workbook(results, uploaded_data, user_classifications,
                        data_type, metadata, ec_store,
                        pa_summary_data=pa_summary_data)
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
