"""Compute SEEA EA Physical Accounts for Lithuanian BBT5.

Produces:
  1. Extent Account — area per habitat type (Ha)
  2. Condition Account — mean EVA scores per habitat type
  3. Supply Table — fish CPUE proxy per habitat type
  4. Combined Excel report
"""
import logging
import os
import sys
from pathlib import Path

import geopandas as gpd
import numpy as np
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Make project root importable so we can read the single source of truth for
# the application version.
_PROJECT_ROOT = Path(__file__).resolve().parent.parent
if str(_PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(_PROJECT_ROOT))
from version import __version__ as APP_VERSION  # noqa: E402

logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")
logger = logging.getLogger(__name__)

EVA_FINAL = os.path.normpath(
    r"C:\Users\DELL\OneDrive - ku.lt\HORIZON_EUROPE\MARBEFES\EVA_FINAL"
)
CORRECTED = os.path.normpath(
    r"C:\Users\DELL\OneDrive - ku.lt\HORIZON_EUROPE\MARBEFES\EVA_FINAL_corrected"
)
OUTPUT_DIR = os.path.normpath(
    os.path.join(os.path.dirname(__file__), "..", "tutorial")
)

# Ensure project root on path for imports
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))


def compute_extent_account():
    """Compute habitat extent (area in Ha) from habitats_final.shp."""
    logger.info("=== EXTENT ACCOUNT ===")
    src = os.path.join(EVA_FINAL, "habitats_EVA", "habitats_final.shp")
    gdf = gpd.read_file(src)
    logger.info("Read %d habitat polygons", len(gdf))

    # Filter out NaN habitat types
    gdf = gdf[gdf["MSFD_broad"].notna()].copy()
    logger.info("  %d polygons with valid habitat type", len(gdf))

    # Compute area in hectares from Shape_Area (m2)
    gdf["area_ha"] = gdf["Shape_Area"] / 10000.0

    # Aggregate by habitat type
    extent = gdf.groupby("MSFD_broad").agg(
        polygon_count=("Shape_Area", "count"),
        total_area_ha=("area_ha", "sum"),
    ).reset_index()
    extent = extent.rename(columns={"MSFD_broad": "Habitat Type"})

    # Add percentage
    total = extent["total_area_ha"].sum()
    extent["pct_of_total"] = (extent["total_area_ha"] / total * 100).round(1)
    extent["total_area_ha"] = extent["total_area_ha"].round(2)

    # Sort by area descending
    extent = extent.sort_values("total_area_ha", ascending=False).reset_index(drop=True)

    # Add total row
    total_row = pd.DataFrame([{
        "Habitat Type": "TOTAL",
        "polygon_count": extent["polygon_count"].sum(),
        "total_area_ha": total.round(2),
        "pct_of_total": 100.0,
    }])
    extent = pd.concat([extent, total_row], ignore_index=True)

    logger.info("Extent Account:")
    logger.info("\n%s", extent.to_string(index=False))
    return extent


def compute_condition_account():
    """Compute ecological condition per habitat type using EVA scores."""
    logger.info("\n=== CONDITION ACCOUNT ===")

    # Read habitats with types
    hab_src = os.path.join(EVA_FINAL, "habitats_EVA", "habitats_final.shp")
    habitats = gpd.read_file(hab_src)
    habitats = habitats[habitats["MSFD_broad"].notna()].copy()

    # Read the corrected combined EVA layer (or original if corrected not available)
    combined_src = os.path.join(CORRECTED, "ALL4EVA_2025_fixed_geometries.gpkg")
    if not os.path.exists(combined_src):
        combined_src = os.path.join(EVA_FINAL, "ALL4EVA_2025_fixed_geometries.gpkg")
    logger.info("Reading combined EVA from %s", combined_src)
    eva = gpd.read_file(combined_src)

    # Spatial join: habitat polygons → EVA grid (using centroids)
    hab_proj = habitats.to_crs(eva.crs) if habitats.crs != eva.crs else habitats
    hab_centroids = hab_proj[["MSFD_broad", "EVA_AQ7", "geometry"]].copy()
    hab_centroids["geometry"] = hab_centroids.geometry.centroid

    # Get EVA scores at habitat centroid locations
    score_cols = [c for c in eva.columns if c not in ["geometry", "Subzone_ID"]]
    joined = gpd.sjoin(hab_centroids, eva[score_cols + ["geometry"]], how="left", predicate="within")

    # Condition indicators per habitat type
    condition_cols = {
        "AQ7_HABITATS": "Habitat Diversity (AQ7)",
        "ZooScore": "Zooplankton Condition",
        "PhytoScore": "Phytoplankton Condition",
        "MaxBenthos": "Benthos Condition (max AQ)",
    }

    available_cols = {k: v for k, v in condition_cols.items() if k in joined.columns}

    condition = joined.groupby("MSFD_broad")[list(available_cols.keys())].agg(
        ["mean", "std", "count"]
    ).round(3)

    # Flatten multi-index columns
    condition.columns = [f"{available_cols.get(c[0], c[0])} ({c[1]})" for c in condition.columns]
    condition = condition.reset_index().rename(columns={"MSFD_broad": "Habitat Type"})

    # Simpler version: just means
    condition_simple = joined.groupby("MSFD_broad")[list(available_cols.keys())].mean().round(3)
    condition_simple = condition_simple.reset_index().rename(columns={"MSFD_broad": "Habitat Type"})
    condition_simple = condition_simple.rename(columns=available_cols)

    # Add EVA 5-class for each condition indicator
    def classify_eva(val):
        if pd.isna(val):
            return "No Data"
        if val <= 1:
            return "Very Low"
        if val <= 2:
            return "Low"
        if val <= 3:
            return "Medium"
        if val <= 4:
            return "High"
        return "Very High"

    for orig_col, nice_name in available_cols.items():
        condition_simple[f"{nice_name} Class"] = condition_simple[nice_name].apply(classify_eva)

    logger.info("Condition Account:")
    logger.info("\n%s", condition_simple.to_string(index=False))
    return condition_simple, condition


def compute_supply_table():
    """Compute ecosystem service supply proxy per habitat type.

    Uses fish CPUE scores as fisheries provisioning proxy.
    Other services noted as 'Data not available'.
    """
    logger.info("\n=== SUPPLY TABLE ===")

    # Read habitats
    hab_src = os.path.join(EVA_FINAL, "habitats_EVA", "habitats_final.shp")
    habitats = gpd.read_file(hab_src)
    habitats = habitats[habitats["MSFD_broad"].notna()].copy()

    # Read fish data
    fish_src = os.path.join(EVA_FINAL, "AllFishAQ.gpkg")
    if os.path.exists(fish_src):
        fish = gpd.read_file(fish_src)
        logger.info("Read %d fish records", len(fish))
    else:
        logger.warning("AllFishAQ.gpkg not found")
        fish = None

    # Read combined layer for other scores
    combined_src = os.path.join(CORRECTED, "ALL4EVA_2025_fixed_geometries.gpkg")
    if not os.path.exists(combined_src):
        combined_src = os.path.join(EVA_FINAL, "ALL4EVA_2025_fixed_geometries.gpkg")
    eva = gpd.read_file(combined_src)

    # Spatial join habitats → EVA grid
    hab_proj = habitats.to_crs(eva.crs) if habitats.crs != eva.crs else habitats
    hab_centroids = hab_proj[["MSFD_broad", "geometry"]].copy()
    hab_centroids["geometry"] = hab_centroids.geometry.centroid

    service_cols = {}
    if "EVA_all_fish" in eva.columns:
        service_cols["EVA_all_fish"] = "Fisheries Provisioning (proxy score)"
    if "ZooScore" in eva.columns:
        service_cols["ZooScore"] = "Food Web Support (zooplankton proxy)"
    if "PhytoScore" in eva.columns:
        service_cols["PhytoScore"] = "Primary Production (phytoplankton proxy)"

    joined = gpd.sjoin(
        hab_centroids,
        eva[list(service_cols.keys()) + ["geometry"]],
        how="left", predicate="within"
    )

    supply = joined.groupby("MSFD_broad")[list(service_cols.keys())].mean().round(3)
    supply = supply.reset_index().rename(columns={"MSFD_broad": "Habitat Type"})
    supply = supply.rename(columns=service_cols)

    # Add placeholder columns for missing services
    supply["Carbon Sequestration (tC/ha/yr)"] = "Data not available"
    supply["Coastal Protection (proxy)"] = "Data not available"
    supply["Tourism & Recreation (visits/yr)"] = "Data not available"
    supply["Nutrient Cycling (proxy)"] = "Data not available"

    logger.info("Supply Table:")
    logger.info("\n%s", supply.to_string(index=False))
    return supply


def write_pa_report(extent, condition_simple, condition_detailed, supply):
    """Write Physical Accounts to a styled Excel workbook."""
    logger.info("\n=== WRITING REPORT ===")
    output_path = os.path.join(OUTPUT_DIR, "MARBEFES_PhysicalAccounts_LithuanianBBT5.xlsx")

    # Style constants
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="006994", end_color="006994", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="D0D0D0"),
        right=Side(style="thin", color="D0D0D0"),
        top=Side(style="thin", color="D0D0D0"),
        bottom=Side(style="thin", color="D0D0D0"),
    )
    alt_fill = PatternFill(start_color="F0F7FA", end_color="F0F7FA", fill_type="solid")

    def style_sheet(ws, start_row=1):
        for cell in ws[start_row]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border
        for row_idx in range(start_row + 1, ws.max_row + 1):
            for cell in ws[row_idx]:
                cell.border = thin_border
                if (row_idx - start_row) % 2 == 0:
                    cell.fill = alt_fill
        for col_idx in range(1, ws.max_column + 1):
            max_len = 0
            for row_idx in range(start_row, min(ws.max_row + 1, start_row + 50)):
                val = ws.cell(row=row_idx, column=col_idx).value
                if val:
                    max_len = max(max_len, len(str(val)))
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 3, 12), 45)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        # Sheet 1: Metadata
        meta = pd.DataFrame({
            "Parameter": [
                "Report Title",
                "Study Area",
                "EAA (Ecosystem Accounting Area)",
                "Accounting Period",
                "Framework",
                "Generated",
                "",
                "Data Sources",
                "Habitats",
                "EVA Scores",
                "Fish Data",
                "",
                "Reference",
                "Funding",
            ],
            "Value": [
                "SEEA EA Physical Accounts — Lithuanian BBT5",
                "Curonian Lagoon and Baltic Sea Coast, Lithuania",
                "Lithuanian EEZ + Territorial Sea + Curonian Lagoon",
                "2017-2023 (monitoring data period)",
                "SEEA EA (UN System of Environmental-Economic Accounting)",
                pd.Timestamp.now().strftime("%Y-%m-%d"),
                "",
                "",
                "HELCOM HUB L3 / EUNIS L2 (Lithuanian EPA, 2019)",
                f"MARBEFES EVA v{APP_VERSION} (corrected dataset, Sept 2025)",
                "ICES BITS + Lithuanian EPA commercial catch (2000-2023)",
                "",
                "Franco A. & Amorim E. (2025) EVA Guidance. MARBEFES WP4.1",
                "EU Horizon Europe MARBEFES Project",
            ],
        })
        meta.to_excel(writer, sheet_name="Metadata", index=False)

        # Sheet 2: Extent Account
        extent.to_excel(writer, sheet_name="Extent Account", index=False, startrow=2)
        ws = writer.sheets["Extent Account"]
        ws.cell(row=1, column=1, value="Ecosystem Extent Account — Lithuanian BBT5")
        ws.cell(row=1, column=1).font = Font(bold=True, size=14, color="006994")
        ws.cell(row=2, column=1, value="Area per habitat type (hectares)")

        # Sheet 3: Condition Account (simple)
        condition_simple.to_excel(writer, sheet_name="Condition Account", index=False, startrow=2)
        ws = writer.sheets["Condition Account"]
        ws.cell(row=1, column=1, value="Ecosystem Condition Account — Lithuanian BBT5")
        ws.cell(row=1, column=1).font = Font(bold=True, size=14, color="006994")
        ws.cell(row=2, column=1, value="Mean EVA condition scores per habitat type (0-5 scale)")

        # Sheet 4: Condition Account (detailed with std and count)
        condition_detailed.to_excel(writer, sheet_name="Condition (Detailed)", index=False, startrow=2)
        ws = writer.sheets["Condition (Detailed)"]
        ws.cell(row=1, column=1, value="Detailed Condition Statistics")
        ws.cell(row=1, column=1).font = Font(bold=True, size=14, color="006994")

        # Sheet 5: Supply Table
        supply.to_excel(writer, sheet_name="Supply Table", index=False, startrow=2)
        ws = writer.sheets["Supply Table"]
        ws.cell(row=1, column=1, value="Ecosystem Service Supply Table — Lithuanian BBT5")
        ws.cell(row=1, column=1).font = Font(bold=True, size=14, color="006994")
        ws.cell(row=2, column=1, value="Proxy scores (0-5) where available; physical units required for full SEEA EA compliance")

        # Sheet 6: Data Gaps
        gaps = pd.DataFrame({
            "Ecosystem Service": [
                "Fisheries Provisioning",
                "Food Web Support",
                "Primary Production",
                "Carbon Sequestration",
                "Coastal Protection",
                "Tourism & Recreation",
                "Nutrient Cycling",
                "Water Purification",
            ],
            "Status": [
                "Proxy available (EVA fish score)",
                "Proxy available (zooplankton score)",
                "Proxy available (phytoplankton score)",
                "NOT AVAILABLE",
                "NOT AVAILABLE",
                "NOT AVAILABLE",
                "NOT AVAILABLE",
                "NOT AVAILABLE",
            ],
            "Data Needed": [
                "Catch statistics in tonnes/year per habitat type",
                "Zooplankton biomass (mg C/m3) per habitat type",
                "Chlorophyll-a or NPP (g C/m2/yr) per habitat type",
                "Organic carbon burial rate (t C/ha/yr)",
                "Wave attenuation coefficients, erosion rates",
                "Visitor counts, willingness-to-pay studies",
                "N/P removal rates (kg/ha/yr)",
                "Pollutant removal efficiency per habitat type",
            ],
            "Potential Source": [
                "Lithuanian EPA fisheries statistics",
                "EPA monitoring + HELCOM COMBINE",
                "Satellite remote sensing (Copernicus Marine)",
                "Sediment core studies, literature values",
                "Hydrodynamic modelling (SHYFEM)",
                "Tourism statistics, recreation surveys",
                "HELCOM PLC data, biogeochemical models",
                "Water quality monitoring, wetland studies",
            ],
        })
        gaps.to_excel(writer, sheet_name="Data Gaps & Next Steps", index=False, startrow=2)
        ws = writer.sheets["Data Gaps & Next Steps"]
        ws.cell(row=1, column=1, value="Data Gaps and Recommendations for Full SEEA EA Compliance")
        ws.cell(row=1, column=1).font = Font(bold=True, size=14, color="006994")

    # Apply styling to all sheets
    wb = openpyxl.load_workbook(output_path)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        start = 3 if sheet_name != "Metadata" else 1
        style_sheet(ws, start_row=start)
        ws.sheet_properties.tabColor = "006994"
    wb.save(output_path)

    logger.info("Report written to %s", output_path)
    return output_path


def main():
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    extent = compute_extent_account()
    condition_simple, condition_detailed = compute_condition_account()
    supply = compute_supply_table()
    path = write_pa_report(extent, condition_simple, condition_detailed, supply)
    logger.info("\nDone. Physical Accounts report: %s", path)


if __name__ == "__main__":
    main()
